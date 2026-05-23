import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Budget DYF 2027"

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_DARK      = "1A0A12"
C_GOLD      = "F5B731"
C_CRIMSON   = "C41E3A"
C_ORANGE    = "E86A17"
C_BORDEAUX  = "6B1D3A"
C_CREAM     = "FFF5E6"
C_TEAL      = "1A6B5A"
C_PURPLE    = "5C2D8F"
C_NAVY      = "1A3A6B"
C_WHITE     = "FFFFFF"
C_GREY_L    = "F7F7F7"
C_GREY_D    = "EEEEEE"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(style="thin", color="DDDDDD"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Largeurs colonnes ─────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16

# ── Ligne 1 : Titre ───────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 52
ws.merge_cells("A1:D1")
c = ws["A1"]
c.value = "DANCE YOUR FOOT · BUDGET ÉDITION 2027"
c.font = Font(bold=True, color=C_GOLD, size=20, name="Calibri")
c.fill = fill(C_DARK)
c.alignment = align("center", "center")

ws.row_dimensions[2].height = 20
ws.merge_cells("A2:D2")
c = ws["A2"]
c.value = "Suisse romande · Première édition · Contact : Sylvain@danceyourfoot.com"
c.font = Font(italic=True, color="AAAAAA", size=10, name="Calibri")
c.fill = fill(C_DARK)
c.alignment = align("center", "center")

ws.row_dimensions[3].height = 8

# ── En-têtes colonnes ─────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 26
for col, h in enumerate(["POSTE", "DÉTAIL / NOTE", "MONTANT (CHF)", "STATUT"], 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(bold=True, color=C_CREAM, size=10, name="Calibri")
    c.fill = fill(C_BORDEAUX)
    c.alignment = align("center", "center")
    c.border = border("medium", C_BORDEAUX)

# ── Sections budget ───────────────────────────────────────────────────────────
sections = [
    {
        "title": "👥  RH & CACHETS",
        "color": C_CRIMSON,
        "rows": [
            ("Coordination",           "Responsable événement — journée complète",                200, "Confirmé"),
            ("DJ",                     "Cachet artiste — tarif estimé marché suisse",              200, "À négocier"),
            ("MC / Speaker",           "Animation publique — cachet",                             200, "À négocier"),
            ("Arbitre",                "Match officiel, briefing inclus",                         200, "Confirmé"),
            ("Photo & vidéo",          "Captation jour J + traitement + montage inclus",         2000, "Confirmé"),
            ("Défraiements équipes",   "4 équipes × 200 CHF (transport, repas)",                  800, "Confirmé"),
        ],
    },
    {
        "title": "🏟️  LIEU & TECHNIQUE",
        "color": C_NAVY,
        "rows": [
            ("Location salle / lieu",  "Espace urbain couvert — estimation milieu de gamme",     2500, "Estimé"),
            ("Sonorisation",           "Enceintes, table de mix, câblage, ingé son",               800, "Estimé"),
            ("Éclairage",              "Jeux de lumière, projecteurs, ambiance scénique",          500, "Estimé"),
            ("Électricité",            "Raccordement technique / générateur si besoin",            200, "Estimé"),
        ],
    },
    {
        "title": "📦  MATÉRIEL & LOGISTIQUE",
        "color": C_BORDEAUX,
        "rows": [
            ("Terrain gonflable",      "Location + installation + personnalisation",              2000, "Confirmé"),
            ("Chasubles personnalisées","30 pièces — crew + équipes",                              400, "Confirmé"),
            ("Ballon personnalisé",    "Ballon officiel aux couleurs DYF",                          60, "Confirmé"),
            ("Cartons public & arbitres","~500 cartons bicolores distribués à l'entrée",           200, "Confirmé"),
            ("Transport & logistique", "Déplacement matériel, manutention",                       300, "Confirmé"),
        ],
    },
    {
        "title": "🛡️  SÉCURITÉ & COUVERTURE",
        "color": C_TEAL,
        "rows": [
            ("Sécurité / Service d'ordre","Agent(s) selon exigences du lieu",                     500, "Estimé"),
            ("Assurance événementielle","RC obligatoire pour location de salle",                   300, "Estimé"),
            ("SUISA",                  "Droits musique live — selon nb spectateurs",               200, "À vérifier"),
        ],
    },
    {
        "title": "📣  COMMUNICATION & RÉCOMPENSES",
        "color": C_ORANGE,
        "rows": [
            ("Communication",          "Affiches, flyers, impression + boost réseaux sociaux",    400, "Estimé"),
            ("Trophées / Prix",        "Récompenses équipes, MVP, mention spéciale DJ",            350, "Estimé"),
            ("Badges & accréditations","Staff, presse, sponsors — impression + cordons",          100, "Estimé"),
        ],
    },
    {
        "title": "🎨  ORGANISATION & ADMINISTRATION",
        "color": C_PURPLE,
        "rows": [
            ("Organisation générale",  "Frais admin, coordination, réunions",                    2000, "Confirmé"),
            ("Création graphique & tech","Identité visuelle, supports, site web",                2000, "Confirmé"),
            ("Divers & imprévus",      "Réserve ~5% du budget total",                             500, "Réserve"),
        ],
    },
]

current_row = 5
subtotal_cells = []

for sec in sections:
    # En-tête section
    ws.row_dimensions[current_row].height = 28
    ws.merge_cells(f"A{current_row}:D{current_row}")
    c = ws.cell(row=current_row, column=1, value=sec["title"])
    c.font = Font(bold=True, color=C_WHITE, size=12, name="Calibri")
    c.fill = fill(sec["color"])
    c.alignment = align("left", "center")
    current_row += 1

    sec_start = current_row
    for i, (poste, detail, montant, statut) in enumerate(sec["rows"]):
        ws.row_dimensions[current_row].height = 20
        row_bg = C_GREY_L if i % 2 == 0 else C_WHITE

        # Couleur statut
        statut_color = {
            "Confirmé":   ("1A6B2A", "E8F5EA"),
            "Estimé":     ("7B5800", "FFF8E1"),
            "À négocier": ("8B3A00", "FFF0E0"),
            "À vérifier": ("555555", "F5F5F5"),
            "Réserve":    ("1A3A6B", "E8F0FB"),
        }.get(statut, ("333333", C_WHITE))

        data = [poste, detail, montant, statut]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=current_row, column=col, value=val)
            c.fill = fill(row_bg)
            c.border = border("thin", "E0E0E0")
            c.alignment = align("right" if col == 3 else "left", "center", wrap=(col == 2))

            if col == 1:
                c.font = Font(bold=True, color="1A1A1A", size=10, name="Calibri")
            elif col == 2:
                c.font = Font(color="555555", size=9, name="Calibri", italic=True)
            elif col == 3:
                c.font = Font(bold=True, color=C_DARK, size=11, name="Calibri")
                c.number_format = '#,##0" CHF"'
            elif col == 4:
                c.font = Font(bold=True, color=statut_color[0], size=9, name="Calibri")
                c.fill = fill(statut_color[1])
                c.alignment = align("center", "center")

        current_row += 1

    # Sous-total
    ws.row_dimensions[current_row].height = 24
    ws.merge_cells(f"A{current_row}:B{current_row}")
    c = ws.cell(row=current_row, column=1,
                value=f"Sous-total — {sec['title'].split('  ')[1]}")
    c.font = Font(bold=True, color=C_WHITE, size=10, name="Calibri")
    c.fill = fill(sec["color"])
    c.alignment = align("right", "center")

    st_ref = f"C{current_row}"
    st = ws.cell(row=current_row, column=3,
                 value=f"=SUM(C{sec_start}:C{current_row-1})")
    st.font = Font(bold=True, color=C_GOLD, size=12, name="Calibri")
    st.fill = fill(C_DARK)
    st.number_format = '#,##0" CHF"'
    st.alignment = align("right", "center")
    subtotal_cells.append(st_ref)

    ws.cell(row=current_row, column=4).fill = fill(C_DARK)
    current_row += 2  # espace entre sections

# ── Total général ─────────────────────────────────────────────────────────────
current_row += 1
ws.row_dimensions[current_row].height = 44
ws.merge_cells(f"A{current_row}:B{current_row}")
c = ws.cell(row=current_row, column=1, value="💰  TOTAL GÉNÉRAL")
c.font = Font(bold=True, color=C_GOLD, size=16, name="Calibri")
c.fill = fill(C_DARK)
c.alignment = align("right", "center")

total = ws.cell(row=current_row, column=3,
                value="=" + "+".join(subtotal_cells))
total.font = Font(bold=True, color=C_GOLD, size=18, name="Calibri")
total.fill = fill(C_DARK)
total.number_format = '#,##0" CHF"'
total.alignment = align("right", "center")

ws.cell(row=current_row, column=4).fill = fill(C_DARK)

# ── Légende statuts ───────────────────────────────────────────────────────────
current_row += 2
ws.row_dimensions[current_row].height = 20
ws.merge_cells(f"A{current_row}:D{current_row}")
c = ws.cell(row=current_row, column=1, value="Légende des statuts :")
c.font = Font(bold=True, color="333333", size=10, name="Calibri")
c.fill = fill("F0F0F0")

current_row += 1
legendes = [
    ("Confirmé",    "Montant validé",              "1A6B2A", "E8F5EA"),
    ("Estimé",      "Montant ajouté par DYF",       "7B5800", "FFF8E1"),
    ("À négocier",  "Tarif à confirmer",            "8B3A00", "FFF0E0"),
    ("À vérifier",  "Dépend du contexte",           "555555", "F5F5F5"),
    ("Réserve",     "Tampon imprévus",              "1A3A6B", "E8F0FB"),
]
for label, desc, fc, bc in legendes:
    ws.row_dimensions[current_row].height = 18
    c1 = ws.cell(row=current_row, column=1, value=label)
    c1.font = Font(bold=True, color=fc, size=9, name="Calibri")
    c1.fill = fill(bc)
    c1.alignment = align("center", "center")
    c2 = ws.cell(row=current_row, column=2, value=desc)
    c2.font = Font(color="555555", size=9, name="Calibri")
    c2.fill = fill("FAFAFA")
    for col in [3, 4]:
        ws.cell(row=current_row, column=col).fill = fill("FAFAFA")
    current_row += 1

ws.freeze_panes = "A5"
ws.sheet_properties.tabColor = C_GOLD

out = "/home/user/danceyourfoot/Budget_DanceYourFoot_2027.xlsx"
wb.save(out)
print(f"Fichier créé : {out}")
